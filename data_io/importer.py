import datetime
from decimal import Decimal, InvalidOperation
import openpyxl
from django.db import transaction
from django.utils import timezone

# Import all models
from income.models import Income
from expenses.models import Expense
from categories.models import Category, Label
from accounts.models import Account, AccountTransfer
from budgets.models import Budget
from goals.models import Goal
from debts.models import Debt, Repayment
from companies.models import CompanyAccount, CompanyIncome, CompanyExpense
from recurrences.models import RecurringTransaction, GeneratedOccurrence

def parse_decimal(val):
    if val is None or val == '':
        return None
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    # String cleaning
    clean_val = str(val).replace('$', '').replace('₹', '').replace('€', '').replace('£', '').replace(',', '').strip()
    try:
        return Decimal(clean_val)
    except InvalidOperation:
        raise ValueError(f"Invalid decimal amount: '{val}'")

def parse_date(val):
    if val is None or val == '':
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    # Try parsing string formats
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%Y/%m/%d'):
        try:
            return datetime.datetime.strptime(str(val).strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date format: '{val}' (Expected YYYY-MM-DD)")

class DataImporter:
    def __init__(self, user, file_obj, is_restore=False):
        self.user = user
        self.file_obj = file_obj
        self.is_restore = is_restore
        
        # Mappings of Excel ID -> DB Object
        self.category_map = {}
        self.label_map = {}
        self.account_map = {}
        self.company_account_map = {}
        self.debt_map = {}
        self.recurring_map = {}
        self.income_map = {}
        self.expense_map = {}
        self.company_income_map = {}
        self.company_expense_map = {}

        # Import statistics
        self.imported_counts = {
            'categories': 0, 'labels': 0, 'accounts': 0, 'income': 0,
            'expenses': 0, 'transfers': 0, 'budgets': 0, 'goals': 0,
            'debts': 0, 'repayments': 0, 'company_accounts': 0,
            'company_income': 0, 'company_expenses': 0,
            'recurring_transactions': 0, 'generated_occurrences': 0
        }
        self.skipped_duplicates = 0
        self.skipped_invalid = 0
        self.errors = [] # List of {"sheet": s, "row": r, "error": e, "data": d}

    def _get_sheet_by_titles(self, wb, titles):
        """Finds a worksheet matching any of the possible titles (case-insensitive)."""
        lower_titles = [t.lower() for t in titles]
        for name in wb.sheetnames:
            if name.lower() in lower_titles:
                return wb[name]
        return None

    def dry_run_validate(self):
        """Performs a structural and basic data validation check on the workbook."""
        validation_errors = []
        validation_warnings = []

        try:
            wb = openpyxl.load_workbook(self.file_obj, read_only=True, data_only=True)
        except Exception as e:
            return False, ["Unable to load Excel workbook. The file might be corrupted or in an invalid format."], []

        # Check required sheets
        required_sheets = ['income', 'expenses', 'categories', 'accounts']
        found_sheets = []
        for req in required_sheets:
            sheet = self._get_sheet_by_titles(wb, [req, req.replace('_', ' ')])
            if sheet:
                found_sheets.append(req)
            else:
                validation_warnings.append(f"Missing sheet recommendation: '{req.title()}' sheet is recommended but not found.")

        if not found_sheets:
            validation_errors.append("No recognizable data sheets found in the Excel file. Verify worksheets (Income, Expenses, Categories, Accounts).")
            return False, validation_errors, validation_warnings

        # Validate headers for the sheets that are present
        specs = self._get_expected_headers()
        for name in wb.sheetnames:
            lower_name = name.lower().replace(' ', '_')
            spec_key = None
            if lower_name in specs:
                spec_key = lower_name
            elif lower_name == 'payment_methods' or lower_name == 'payment_method':
                spec_key = 'payment_methods'

            if spec_key:
                expected_headers = specs[spec_key]
                ws = wb[name]
                
                # Get the first 5 rows to analyze where headers/banners are
                rows_preview = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
                
                # Filter out completely empty rows (where all cells are None or empty string)
                non_empty_rows = []
                for r_idx, r in enumerate(rows_preview, 1):
                    if any(cell is not None and str(cell).strip() != '' for cell in r):
                        non_empty_rows.append((r_idx, r))
                
                if not non_empty_rows:
                    validation_errors.append(f"Sheet '{name}' is empty and lacks header row.")
                    continue
                
                first_content_row_idx, first_content_row = non_empty_rows[0]
                headers = [str(h).strip().lower() for h in first_content_row if h is not None]
                is_banner = len(headers) == 1 or (len(headers) > 1 and "report" in headers[0])
                
                if is_banner:
                    # Header should be in the next non-empty row
                    if len(non_empty_rows) > 1:
                        _, second_content_row = non_empty_rows[1]
                        headers = [str(h).strip().lower() for h in second_content_row if h is not None]
                    else:
                        validation_errors.append(f"Sheet '{name}' contains a banner, but actual headers are missing.")
                        continue

                expected_lower = [h.lower() for h in expected_headers]
                missing_cols = []
                for exp in expected_lower:
                    # Allow match if it contains a substring or matches exactly
                    matched = False
                    for h in headers:
                        if exp in h or h in exp:
                            matched = True
                            break
                    if not matched:
                        missing_cols.append(exp)

                if missing_cols:
                    validation_errors.append(f"Sheet '{name}' is missing required columns: {', '.join(missing_cols)}.")

        is_valid = len(validation_errors) == 0
        return is_valid, validation_errors, validation_warnings

    def _get_expected_headers(self):
        """Expected lowercase substrings/headers for sheet columns."""
        return {
            'categories': ['category id', 'name'],
            'labels': ['label id', 'name'],
            'accounts': ['account id', 'name', 'account type'],
            'payment_methods': ['payment method id', 'name', 'account type'],
            'income': ['amount', 'source', 'date'],
            'expenses': ['name', 'amount', 'date'],
            'transfers': ['amount', 'from account', 'to account', 'date'],
            'budgets': ['category', 'budget amount', 'month', 'year'],
            'goals': ['name', 'target amount'],
            'debts': ['person name', 'debt type', 'amount', 'date'],
            'repayments': ['debt', 'amount', 'date'],
            'company_accounts': ['name', 'opening balance'],
            'company_income': ['company account', 'amount', 'source', 'date'],
            'company_expenses': ['company account', 'name', 'amount', 'date'],
            'recurring_transactions': ['name', 'type', 'amount', 'frequency', 'start date'],
            'generated_occurrences': ['recurring', 'occurrence date']
        }

    def import_data(self):
        """Runs the import within a transaction.atomic block."""
        try:
            wb = openpyxl.load_workbook(self.file_obj, data_only=True)
        except Exception as e:
            self.errors.append({"sheet": "Workbook", "row": 0, "error": f"Error loading workbook: {str(e)}", "data": ""})
            return False

        # Run in atomic transaction
        with transaction.atomic():
            # 1. Categories
            self._import_categories(wb)
            # 2. Labels
            self._import_labels(wb)
            # 3. Accounts & Payment Methods
            self._import_accounts(wb)
            # 4. Incomes & Expenses & Transfers
            self._import_incomes(wb)
            self._import_expenses(wb)
            self._import_transfers(wb)
            # 5. Budgets & Goals
            self._import_budgets(wb)
            self._import_goals(wb)
            # 6. Company Accounts, Company Income, Company Expense
            self._import_company_accounts(wb)
            self._import_company_income(wb)
            self._import_company_expenses(wb)
            # 7. Debts & Repayments
            self._import_debts(wb)
            self._import_repayments(wb)
            # 8. Recurring Transactions & Occurrences
            self._import_recurring_transactions(wb)
            self._import_occurrences(wb)

        return len(self.errors) == 0

    def _get_sheet_rows(self, wb, sheet_slug):
        """Helper to get headers list and rows values list from sheet, ignoring banners."""
        ws = self._get_sheet_by_titles(wb, [sheet_slug, sheet_slug.replace('_', ' ')])
        if not ws:
            return None, []

        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return None, []

        # Find the first non-empty row index
        non_empty_rows = []
        for r_idx, r in enumerate(all_rows):
            if any(cell is not None and str(cell).strip() != '' for cell in r):
                non_empty_rows.append((r_idx, r))

        if not non_empty_rows:
            return None, []

        first_content_idx, first_content_row = non_empty_rows[0]
        headers = [str(h).strip().lower() if h is not None else '' for h in first_content_row]
        
        # Check if first content row is a banner (only one non-empty value or "report" in the first value)
        headers_filtered = [h for h in headers if h != '']
        is_banner = len(headers_filtered) == 1 or (len(headers_filtered) > 1 and "report" in headers_filtered[0])
        
        start_idx = first_content_idx + 1
        
        if is_banner:
            # Look at the next non-empty row for actual headers
            second_content_idx = None
            for idx, r in non_empty_rows:
                if idx > first_content_idx:
                    second_content_idx = idx
                    break
            
            if second_content_idx is not None:
                headers = [str(h).strip().lower() if h is not None else '' for h in all_rows[second_content_idx]]
                start_idx = second_content_idx + 1
            else:
                return None, []

        # Parse rows starting from start_idx
        data_rows = []
        for r_num, row in enumerate(all_rows[start_idx:], start_idx + 1):
            # Check if row is empty
            if any(cell is not None and str(cell).strip() != '' for cell in row):
                data_rows.append((r_num, row))

        return headers, data_rows

    def _find_col_idx(self, headers, col_slugs):
        """Finds header index matches."""
        for slug in col_slugs:
            for idx, h in enumerate(headers):
                if slug in h:
                    return idx
        return -1

    def _import_categories(self, wb):
        headers, rows = self._get_sheet_rows(wb, 'categories')
        if not headers:
            return

        id_idx = self._find_col_idx(headers, ['category id', 'id'])
        name_idx = self._find_col_idx(headers, ['name'])
        desc_idx = self._find_col_idx(headers, ['description', 'desc'])

        if name_idx == -1:
            return

        for r_num, row in rows:
            name = str(row[name_idx]).strip() if row[name_idx] else None
            excel_id = row[id_idx] if id_idx != -1 else None
            desc = row[desc_idx] if desc_idx != -1 else None

            if not name:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Categories", "row": r_num, "error": "Category Name is required.", "data": str(row)})
                continue

            try:
                # get or create category (prevent name duplicates per user)
                cat, created = Category.objects.get_or_create(
                    user=self.user,
                    name=name,
                    defaults={'description': desc}
                )
                if created:
                    self.imported_counts['categories'] += 1
                else:
                    self.skipped_duplicates += 1

                if excel_id is not None:
                    self.category_map[excel_id] = cat
            except Exception as e:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Categories", "row": r_num, "error": str(e), "data": str(row)})

    def _import_labels(self, wb):
        headers, rows = self._get_sheet_rows(wb, 'labels')
        if not headers:
            return

        id_idx = self._find_col_idx(headers, ['label id', 'id'])
        name_idx = self._find_col_idx(headers, ['name'])
        color_idx = self._find_col_idx(headers, ['color'])

        if name_idx == -1:
            return

        for r_num, row in rows:
            name = str(row[name_idx]).strip() if row[name_idx] else None
            excel_id = row[id_idx] if id_idx != -1 else None
            color = str(row[color_idx]).strip() if color_idx != -1 and row[color_idx] else '#6366f1'

            if not name:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Labels", "row": r_num, "error": "Label Name is required.", "data": str(row)})
                continue

            try:
                label, created = Label.objects.get_or_create(
                    user=self.user,
                    name=name,
                    defaults={'color': color}
                )
                if created:
                    self.imported_counts['labels'] += 1
                else:
                    self.skipped_duplicates += 1

                if excel_id is not None:
                    self.label_map[excel_id] = label
            except Exception as e:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Labels", "row": r_num, "error": str(e), "data": str(row)})

    def _import_accounts(self, wb):
        # We check both 'accounts' and 'payment_methods' sheets
        for sheet_slug in ['accounts', 'payment_methods']:
            headers, rows = self._get_sheet_rows(wb, sheet_slug)
            if not headers:
                continue

            id_idx = self._find_col_idx(headers, ['id', 'account id', 'payment method id'])
            name_idx = self._find_col_idx(headers, ['name'])
            type_idx = self._find_col_idx(headers, ['type', 'account type'])
            init_idx = self._find_col_idx(headers, ['initial balance', 'initial'])
            min_idx = self._find_col_idx(headers, ['minimum balance', 'minimum'])
            status_idx = self._find_col_idx(headers, ['status'])
            notes_idx = self._find_col_idx(headers, ['notes', 'note'])

            if name_idx == -1:
                continue

            for r_num, row in rows:
                name = str(row[name_idx]).strip() if row[name_idx] else None
                excel_id = row[id_idx] if id_idx != -1 else None
                acc_type = str(row[type_idx]).strip() if type_idx != -1 and row[type_idx] else 'Cash'
                
                # Clean account type to match choices
                valid_types = [t[0] for t in Account.ACCOUNT_TYPES]
                matched_type = 'Cash'
                for vt in valid_types:
                    if vt.lower() == acc_type.lower() or vt.lower().replace(' ', '') == acc_type.lower().replace(' ', ''):
                        matched_type = vt
                        break

                initial_bal = parse_decimal(row[init_idx]) if init_idx != -1 and row[init_idx] is not None else Decimal('0.00')
                minimum_bal = parse_decimal(row[min_idx]) if min_idx != -1 and row[min_idx] is not None else Decimal('0.00')
                status = str(row[status_idx]).strip().upper() if status_idx != -1 and row[status_idx] else 'ACTIVE'
                if status not in ['ACTIVE', 'INACTIVE', 'CLOSED']:
                    status = 'ACTIVE'
                notes = row[notes_idx] if notes_idx != -1 else None

                if not name:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": sheet_slug.title(), "row": r_num, "error": "Account Name is required.", "data": str(row)})
                    continue

                try:
                    # Retrieve or create account
                    acc, created = Account.objects.get_or_create(
                        user=self.user,
                        name=name,
                        defaults={
                            'account_type': matched_type,
                            'initial_balance': initial_bal,
                            'minimum_balance': minimum_bal,
                            'status': status,
                            'notes': notes
                        }
                    )
                    
                    # Ensure minimum balance threshold edit flag is active during imports to bypass validators if needed
                    if not created:
                        # Skip adjusting initial balance, but map the ID
                        self.skipped_duplicates += 1
                    else:
                        self.imported_counts['accounts'] += 1

                    if excel_id is not None:
                        self.account_map[excel_id] = acc
                except Exception as e:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": sheet_slug.title(), "row": r_num, "error": str(e), "data": str(row)})

    def _get_resolved_account(self, excel_id, name):
        """Lookup account by mapped ID, or by name, or return fallback/default account."""
        if excel_id is not None and excel_id in self.account_map:
            return self.account_map[excel_id]
        if name:
            acc = Account.objects.filter(user=self.user, name=name).first()
            if acc:
                return acc
        
        # Fallback default Cash account
        acc, _ = Account.objects.get_or_create(
            user=self.user,
            name='Cash',
            defaults={'account_type': 'Cash', 'initial_balance': Decimal('0.00')}
        )
        return acc

    def _get_resolved_category(self, excel_id, name):
        """Lookup category by mapped ID, or by name, or create automatically if missing."""
        if excel_id is not None and excel_id in self.category_map:
            return self.category_map[excel_id]
        if name:
            name_clean = name.strip()
            if name_clean:
                cat, created = Category.objects.get_or_create(
                    user=self.user,
                    name=name_clean,
                    defaults={'description': 'Auto-created during import.'}
                )
                if created:
                    self.imported_counts['categories'] += 1
                return cat
        return None

    def _import_incomes(self, wb):
        headers, rows = self._get_sheet_rows(wb, 'income')
        if not headers:
            return

        id_idx = self._find_col_idx(headers, ['id', 'income id'])
        amount_idx = self._find_col_idx(headers, ['amount'])
        source_idx = self._find_col_idx(headers, ['source'])
        date_idx = self._find_col_idx(headers, ['date'])
        acc_id_idx = self._find_col_idx(headers, ['account id'])
        acc_name_idx = self._find_col_idx(headers, ['account name', 'account'])
        label_idx = self._find_col_idx(headers, ['labels'])
        desc_idx = self._find_col_idx(headers, ['description', 'note'])

        if amount_idx == -1 or date_idx == -1 or source_idx == -1:
            return

        for r_num, row in rows:
            excel_id = row[id_idx] if id_idx != -1 else None
            
            try:
                amount = parse_decimal(row[amount_idx])
                date = parse_date(row[date_idx])
                source = str(row[source_idx]).strip() if row[source_idx] else 'Other'
                
                # Standardize source choices
                valid_sources = [s[0] for s in Income.SOURCE_CHOICES]
                if source not in valid_sources:
                    source = 'Other'

                acc_id = row[acc_id_idx] if acc_id_idx != -1 else None
                acc_name = str(row[acc_name_idx]).strip() if acc_name_idx != -1 and row[acc_name_idx] else None
                db_account = self._get_resolved_account(acc_id, acc_name)

                description = row[desc_idx] if desc_idx != -1 else None
                labels_raw = str(row[label_idx]).strip() if label_idx != -1 and row[label_idx] else ''

                if amount is None or amount <= 0:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Income", "row": r_num, "error": "Income amount must be greater than zero.", "data": str(row)})
                    continue
                if not date:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Income", "row": r_num, "error": "Income date is required.", "data": str(row)})
                    continue

                # Check duplicate
                duplicate = Income.objects.filter(
                    user=self.user,
                    amount=amount,
                    source=source,
                    date=date,
                    account=db_account,
                    description=description
                ).first()

                if duplicate:
                    self.skipped_duplicates += 1
                    if excel_id is not None:
                        self.income_map[excel_id] = duplicate
                    continue

                # Create income
                income = Income.objects.create(
                    user=self.user,
                    account=db_account,
                    amount=amount,
                    source=source,
                    date=date,
                    description=description
                )

                # Process labels
                if labels_raw:
                    for l_name in labels_raw.split(','):
                        l_name_clean = l_name.strip()
                        if l_name_clean:
                            lbl, created = Label.objects.get_or_create(user=self.user, name=l_name_clean)
                            if created:
                                self.imported_counts['labels'] += 1
                            income.labels.add(lbl)

                self.imported_counts['income'] += 1
                if excel_id is not None:
                    self.income_map[excel_id] = income
            except Exception as e:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Income", "row": r_num, "error": str(e), "data": str(row)})

    def _import_expenses(self, wb):
        headers, rows = self._get_sheet_rows(wb, 'expenses')
        if not headers:
            return

        id_idx = self._find_col_idx(headers, ['id', 'expense id'])
        name_idx = self._find_col_idx(headers, ['name'])
        amount_idx = self._find_col_idx(headers, ['amount'])
        date_idx = self._find_col_idx(headers, ['date'])
        acc_id_idx = self._find_col_idx(headers, ['account id'])
        acc_name_idx = self._find_col_idx(headers, ['account name', 'account'])
        cat_id_idx = self._find_col_idx(headers, ['category id'])
        cat_name_idx = self._find_col_idx(headers, ['category name', 'category'])
        label_idx = self._find_col_idx(headers, ['labels'])
        desc_idx = self._find_col_idx(headers, ['description', 'note'])

        if name_idx == -1 or amount_idx == -1 or date_idx == -1:
            return

        for r_num, row in rows:
            excel_id = row[id_idx] if id_idx != -1 else None

            try:
                name = str(row[name_idx]).strip() if row[name_idx] else None
                amount = parse_decimal(row[amount_idx])
                date = parse_date(row[date_idx])

                acc_id = row[acc_id_idx] if acc_id_idx != -1 else None
                acc_name = str(row[acc_name_idx]).strip() if acc_name_idx != -1 and row[acc_name_idx] else None
                db_account = self._get_resolved_account(acc_id, acc_name)

                cat_id = row[cat_id_idx] if cat_id_idx != -1 else None
                cat_name = str(row[cat_name_idx]).strip() if cat_name_idx != -1 and row[cat_name_idx] else None
                db_category = self._get_resolved_category(cat_id, cat_name)

                description = row[desc_idx] if desc_idx != -1 else None
                labels_raw = str(row[label_idx]).strip() if label_idx != -1 and row[label_idx] else ''

                if not name:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Expenses", "row": r_num, "error": "Expense name is required.", "data": str(row)})
                    continue
                if amount is None or amount <= 0:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Expenses", "row": r_num, "error": "Expense amount must be greater than zero.", "data": str(row)})
                    continue
                if not date:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Expenses", "row": r_num, "error": "Expense date is required.", "data": str(row)})
                    continue

                # Check duplicate
                duplicate = Expense.objects.filter(
                    user=self.user,
                    name=name,
                    amount=amount,
                    date=date,
                    account=db_account,
                    category=db_category,
                    description=description
                ).first()

                if duplicate:
                    self.skipped_duplicates += 1
                    if excel_id is not None:
                        self.expense_map[excel_id] = duplicate
                    continue

                expense = Expense.objects.create(
                    user=self.user,
                    name=name,
                    amount=amount,
                    date=date,
                    account=db_account,
                    category=db_category,
                    description=description
                )

                if labels_raw:
                    for l_name in labels_raw.split(','):
                        l_name_clean = l_name.strip()
                        if l_name_clean:
                            lbl, created = Label.objects.get_or_create(user=self.user, name=l_name_clean)
                            if created:
                                self.imported_counts['labels'] += 1
                            expense.labels.add(lbl)

                self.imported_counts['expenses'] += 1
                if excel_id is not None:
                    self.expense_map[excel_id] = expense
            except Exception as e:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Expenses", "row": r_num, "error": str(e), "data": str(row)})

    def _import_transfers(self, wb):
        headers, rows = self._get_sheet_rows(wb, 'transfers')
        if not headers:
            return

        amount_idx = self._find_col_idx(headers, ['amount'])
        from_id_idx = self._find_col_idx(headers, ['from account id'])
        from_name_idx = self._find_col_idx(headers, ['from account name', 'from account', 'from'])
        to_id_idx = self._find_col_idx(headers, ['to account id'])
        to_name_idx = self._find_col_idx(headers, ['to account name', 'to account', 'to'])
        date_idx = self._find_col_idx(headers, ['date', 'transfer date'])
        note_idx = self._find_col_idx(headers, ['note', 'description'])

        if amount_idx == -1 or date_idx == -1:
            return

        for r_num, row in rows:
            try:
                amount = parse_decimal(row[amount_idx])
                date = parse_date(row[date_idx])

                from_id = row[from_id_idx] if from_id_idx != -1 else None
                from_name = str(row[from_name_idx]).strip() if from_name_idx != -1 and row[from_name_idx] else None
                db_from = self._get_resolved_account(from_id, from_name)

                to_id = row[to_id_idx] if to_id_idx != -1 else None
                to_name = str(row[to_name_idx]).strip() if to_name_idx != -1 and row[to_name_idx] else None
                db_to = self._get_resolved_account(to_id, to_name)

                note = row[note_idx] if note_idx != -1 else None

                if amount is None or amount <= 0:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Transfers", "row": r_num, "error": "Transfer amount must be positive.", "data": str(row)})
                    continue
                if db_from == db_to:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Transfers", "row": r_num, "error": "From and To accounts cannot be the same.", "data": str(row)})
                    continue

                # Check duplicate
                duplicate = AccountTransfer.objects.filter(
                    user=self.user,
                    amount=amount,
                    from_account=db_from,
                    to_account=db_to,
                    transfer_date=date,
                    note=note
                ).exists()

                if duplicate:
                    self.skipped_duplicates += 1
                    continue

                AccountTransfer.objects.create(
                    user=self.user,
                    from_account=db_from,
                    to_account=db_to,
                    amount=amount,
                    transfer_date=date,
                    note=note
                )
                self.imported_counts['transfers'] += 1
            except Exception as e:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Transfers", "row": r_num, "error": str(e), "data": str(row)})

    def _import_budgets(self, wb):
        headers, rows = self._get_sheet_rows(wb, 'budgets')
        if not headers:
            return

        cat_id_idx = self._find_col_idx(headers, ['category id'])
        cat_name_idx = self._find_col_idx(headers, ['category name', 'category'])
        amount_idx = self._find_col_idx(headers, ['amount', 'budget amount'])
        month_idx = self._find_col_idx(headers, ['month'])
        year_idx = self._find_col_idx(headers, ['year'])
        notes_idx = self._find_col_idx(headers, ['notes', 'note'])
        active_idx = self._find_col_idx(headers, ['active', 'is active'])

        if amount_idx == -1 or month_idx == -1 or year_idx == -1:
            return

        for r_num, row in rows:
            try:
                cat_id = row[cat_id_idx] if cat_id_idx != -1 else None
                cat_name = str(row[cat_name_idx]).strip() if cat_name_idx != -1 and row[cat_name_idx] else None
                db_category = self._get_resolved_category(cat_id, cat_name)

                if not db_category:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Budgets", "row": r_num, "error": "Budget Category is required.", "data": str(row)})
                    continue

                amount = parse_decimal(row[amount_idx])
                month = int(row[month_idx])
                year = int(row[year_idx])
                notes = row[notes_idx] if notes_idx != -1 else None
                is_active = True
                if active_idx != -1 and row[active_idx] is not None:
                    is_active = str(row[active_idx]).strip().lower() in ('true', '1', 'yes', 't')

                if amount is None or amount <= 0:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Budgets", "row": r_num, "error": "Budget amount must be positive.", "data": str(row)})
                    continue
                if not (1 <= month <= 12):
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Budgets", "row": r_num, "error": "Month must be between 1 and 12.", "data": str(row)})
                    continue

                # Check duplicate
                duplicate = Budget.objects.filter(
                    category=db_category,
                    month=month,
                    year=year
                ).exists()

                if duplicate:
                    self.skipped_duplicates += 1
                    continue

                Budget.objects.create(
                    user=self.user,
                    category=db_category,
                    budget_amount=amount,
                    month=month,
                    year=year,
                    notes=notes,
                    is_active=is_active
                )
                self.imported_counts['budgets'] += 1
            except Exception as e:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Budgets", "row": r_num, "error": str(e), "data": str(row)})

    def _import_goals(self, wb):
        headers, rows = self._get_sheet_rows(wb, 'goals')
        if not headers:
            return

        name_idx = self._find_col_idx(headers, ['name'])
        target_idx = self._find_col_idx(headers, ['target amount'])
        current_idx = self._find_col_idx(headers, ['current amount'])
        date_idx = self._find_col_idx(headers, ['target date'])
        desc_idx = self._find_col_idx(headers, ['description', 'notes', 'note'])

        if name_idx == -1 or target_idx == -1:
            return

        for r_num, row in rows:
            try:
                name = str(row[name_idx]).strip() if row[name_idx] else None
                target_amount = parse_decimal(row[target_idx])
                current_amount = parse_decimal(row[current_idx]) if current_idx != -1 and row[current_idx] is not None else Decimal('0.00')
                target_date = parse_date(row[date_idx]) if date_idx != -1 else None
                description = row[desc_idx] if desc_idx != -1 else None

                if not name:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Goals", "row": r_num, "error": "Goal name is required.", "data": str(row)})
                    continue
                if target_amount is None or target_amount <= 0:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Goals", "row": r_num, "error": "Goal target amount must be positive.", "data": str(row)})
                    continue

                # Check duplicate
                duplicate = Goal.objects.filter(
                    user=self.user,
                    name=name,
                    target_amount=target_amount,
                    target_date=target_date
                ).exists()

                if duplicate:
                    self.skipped_duplicates += 1
                    continue

                Goal.objects.create(
                    user=self.user,
                    name=name,
                    target_amount=target_amount,
                    current_amount=current_amount,
                    target_date=target_date,
                    description=description
                )
                self.imported_counts['goals'] += 1
            except Exception as e:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Goals", "row": r_num, "error": str(e), "data": str(row)})

    def _import_debts(self, wb):
        headers, rows = self._get_sheet_rows(wb, 'debts')
        if not headers:
            return

        id_idx = self._find_col_idx(headers, ['id', 'debt id'])
        person_idx = self._find_col_idx(headers, ['person name', 'person'])
        type_idx = self._find_col_idx(headers, ['type', 'debt type'])
        amount_idx = self._find_col_idx(headers, ['amount'])
        date_idx = self._find_col_idx(headers, ['date'])
        due_idx = self._find_col_idx(headers, ['due date'])
        notes_idx = self._find_col_idx(headers, ['notes', 'note'])
        status_idx = self._find_col_idx(headers, ['status'])

        if person_idx == -1 or type_idx == -1 or amount_idx == -1 or date_idx == -1:
            return

        for r_num, row in rows:
            excel_id = row[id_idx] if id_idx != -1 else None

            try:
                person_name = str(row[person_idx]).strip() if row[person_idx] else None
                debt_type = str(row[type_idx]).strip() if row[type_idx] else 'Borrowed'
                
                # Check choices
                if 'borrow' in debt_type.lower() or debt_type.lower() == 'borrowed':
                    debt_type = 'Borrowed'
                else:
                    debt_type = 'Lent'

                amount = parse_decimal(row[amount_idx])
                date = parse_date(row[date_idx])
                due_date = parse_date(row[due_idx]) if due_idx != -1 else None
                notes = row[notes_idx] if notes_idx != -1 else None
                status = str(row[status_idx]).strip().capitalize() if status_idx != -1 and row[status_idx] else 'Active'
                if status not in ['Active', 'Settled']:
                    status = 'Active'

                if not person_name:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Debts", "row": r_num, "error": "Person name is required.", "data": str(row)})
                    continue
                if amount is None or amount <= 0:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Debts", "row": r_num, "error": "Debt amount must be positive.", "data": str(row)})
                    continue
                if not date:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Debts", "row": r_num, "error": "Debt date is required.", "data": str(row)})
                    continue

                # Check duplicate
                duplicate = Debt.objects.filter(
                    user=self.user,
                    person_name=person_name,
                    debt_type=debt_type,
                    amount=amount,
                    date=date
                ).first()

                if duplicate:
                    self.skipped_duplicates += 1
                    if excel_id is not None:
                        self.debt_map[excel_id] = duplicate
                    continue

                debt = Debt.objects.create(
                    user=self.user,
                    person_name=person_name,
                    debt_type=debt_type,
                    amount=amount,
                    date=date,
                    due_date=due_date,
                    notes=notes,
                    status=status
                )
                self.imported_counts['debts'] += 1
                if excel_id is not None:
                    self.debt_map[excel_id] = debt
            except Exception as e:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Debts", "row": r_num, "error": str(e), "data": str(row)})

    def _import_repayments(self, wb):
        headers, rows = self._get_sheet_rows(wb, 'repayments')
        if not headers:
            return

        debt_id_idx = self._find_col_idx(headers, ['debt id'])
        debt_name_idx = self._find_col_idx(headers, ['debt person name', 'debt'])
        amount_idx = self._find_col_idx(headers, ['amount'])
        date_idx = self._find_col_idx(headers, ['date'])
        notes_idx = self._find_col_idx(headers, ['notes', 'note'])

        if amount_idx == -1 or date_idx == -1:
            return

        for r_num, row in rows:
            try:
                debt_id = row[debt_id_idx] if debt_id_idx != -1 else None
                debt_name = str(row[debt_name_idx]).strip() if debt_name_idx != -1 and row[debt_name_idx] else None
                
                db_debt = None
                if debt_id is not None and debt_id in self.debt_map:
                    db_debt = self.debt_map[debt_id]
                elif debt_name:
                    db_debt = Debt.objects.filter(user=self.user, person_name=debt_name).first()

                if not db_debt:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Repayments", "row": r_num, "error": f"Referenced Debt not found (ID: {debt_id}, Name: {debt_name}).", "data": str(row)})
                    continue

                amount = parse_decimal(row[amount_idx])
                date = parse_date(row[date_idx])
                notes = row[notes_idx] if notes_idx != -1 else None

                if amount is None or amount <= 0:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Repayments", "row": r_num, "error": "Repayment amount must be positive.", "data": str(row)})
                    continue
                if not date:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Repayments", "row": r_num, "error": "Repayment date is required.", "data": str(row)})
                    continue

                # Check duplicate
                duplicate = Repayment.objects.filter(
                    debt=db_debt,
                    amount=amount,
                    date=date
                ).exists()

                if duplicate:
                    self.skipped_duplicates += 1
                    continue

                Repayment.objects.create(
                    debt=db_debt,
                    amount=amount,
                    date=date,
                    notes=notes
                )
                self.imported_counts['repayments'] += 1
            except Exception as e:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Repayments", "row": r_num, "error": str(e), "data": str(row)})

    def _import_company_accounts(self, wb):
        headers, rows = self._get_sheet_rows(wb, 'company_accounts')
        if not headers:
            return

        id_idx = self._find_col_idx(headers, ['id', 'company account id'])
        name_idx = self._find_col_idx(headers, ['name'])
        desc_idx = self._find_col_idx(headers, ['description', 'desc'])
        open_idx = self._find_col_idx(headers, ['opening balance', 'opening'])
        date_idx = self._find_col_idx(headers, ['created date', 'date'])
        status_idx = self._find_col_idx(headers, ['status'])

        if name_idx == -1:
            return

        for r_num, row in rows:
            excel_id = row[id_idx] if id_idx != -1 else None

            try:
                name = str(row[name_idx]).strip() if row[name_idx] else None
                desc = row[desc_idx] if desc_idx != -1 else None
                opening_balance = parse_decimal(row[open_idx]) if open_idx != -1 and row[open_idx] is not None else Decimal('0.00')
                created_date = parse_date(row[date_idx]) if date_idx != -1 and row[date_idx] is not None else timezone.localdate()
                status = str(row[status_idx]).strip().upper() if status_idx != -1 and row[status_idx] else 'ACTIVE'
                if status not in ['ACTIVE', 'INACTIVE']:
                    status = 'ACTIVE'

                if not name:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Company Accounts", "row": r_num, "error": "Company Account Name is required.", "data": str(row)})
                    continue

                # Get or create Company Account
                comp, created = CompanyAccount.objects.get_or_create(
                    user=self.user,
                    name=name,
                    defaults={
                        'description': desc,
                        'opening_balance': opening_balance,
                        'created_date': created_date,
                        'status': status
                    }
                )
                if created:
                    self.imported_counts['company_accounts'] += 1
                else:
                    self.skipped_duplicates += 1

                if excel_id is not None:
                    self.company_account_map[excel_id] = comp
            except Exception as e:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Company Accounts", "row": r_num, "error": str(e), "data": str(row)})

    def _get_resolved_company_account(self, excel_id, name):
        """Resolves company account by mapped ID or name, or creates default."""
        if excel_id is not None and excel_id in self.company_account_map:
            return self.company_account_map[excel_id]
        if name:
            comp = CompanyAccount.objects.filter(user=self.user, name=name).first()
            if comp:
                return comp
        # Create default
        comp, _ = CompanyAccount.objects.get_or_create(
            user=self.user,
            name='Default Company',
            defaults={'opening_balance': Decimal('0.00')}
        )
        return comp

    def _import_company_income(self, wb):
        headers, rows = self._get_sheet_rows(wb, 'company_income')
        if not headers:
            return

        id_idx = self._find_col_idx(headers, ['id', 'company income id'])
        comp_id_idx = self._find_col_idx(headers, ['company account id'])
        comp_name_idx = self._find_col_idx(headers, ['company account name', 'company account', 'company'])
        amount_idx = self._find_col_idx(headers, ['amount'])
        source_idx = self._find_col_idx(headers, ['source'])
        date_idx = self._find_col_idx(headers, ['date'])
        desc_idx = self._find_col_idx(headers, ['description', 'note'])

        if amount_idx == -1 or date_idx == -1 or source_idx == -1:
            return

        for r_num, row in rows:
            excel_id = row[id_idx] if id_idx != -1 else None

            try:
                comp_id = row[comp_id_idx] if comp_id_idx != -1 else None
                comp_name = str(row[comp_name_idx]).strip() if comp_name_idx != -1 and row[comp_name_idx] else None
                db_comp = self._get_resolved_company_account(comp_id, comp_name)

                amount = parse_decimal(row[amount_idx])
                source = str(row[source_idx]).strip() if row[source_idx] else 'Other'
                date = parse_date(row[date_idx])
                description = row[desc_idx] if desc_idx != -1 else None

                if amount is None or amount <= 0:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Company Income", "row": r_num, "error": "Amount must be positive.", "data": str(row)})
                    continue
                if not date:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Company Income", "row": r_num, "error": "Date is required.", "data": str(row)})
                    continue

                # Check duplicate
                duplicate = CompanyIncome.objects.filter(
                    company_account=db_comp,
                    amount=amount,
                    source=source,
                    date=date
                ).first()

                if duplicate:
                    self.skipped_duplicates += 1
                    if excel_id is not None:
                        self.company_income_map[excel_id] = duplicate
                    continue

                comp_inc = CompanyIncome.objects.create(
                    company_account=db_comp,
                    amount=amount,
                    source=source,
                    date=date,
                    description=description
                )
                self.imported_counts['company_income'] += 1
                if excel_id is not None:
                    self.company_income_map[excel_id] = comp_inc
            except Exception as e:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Company Income", "row": r_num, "error": str(e), "data": str(row)})

    def _import_company_expenses(self, wb):
        headers, rows = self._get_sheet_rows(wb, 'company_expenses')
        if not headers:
            return

        id_idx = self._find_col_idx(headers, ['id', 'company expense id'])
        comp_id_idx = self._find_col_idx(headers, ['company account id'])
        comp_name_idx = self._find_col_idx(headers, ['company account name', 'company account', 'company'])
        name_idx = self._find_col_idx(headers, ['name'])
        amount_idx = self._find_col_idx(headers, ['amount'])
        cat_id_idx = self._find_col_idx(headers, ['category id'])
        cat_name_idx = self._find_col_idx(headers, ['category name', 'category'])
        date_idx = self._find_col_idx(headers, ['date'])
        desc_idx = self._find_col_idx(headers, ['description', 'note'])

        if name_idx == -1 or amount_idx == -1 or date_idx == -1:
            return

        for r_num, row in rows:
            excel_id = row[id_idx] if id_idx != -1 else None

            try:
                comp_id = row[comp_id_idx] if comp_id_idx != -1 else None
                comp_name = str(row[comp_name_idx]).strip() if comp_name_idx != -1 and row[comp_name_idx] else None
                db_comp = self._get_resolved_company_account(comp_id, comp_name)

                name = str(row[name_idx]).strip() if row[name_idx] else None
                amount = parse_decimal(row[amount_idx])
                date = parse_date(row[date_idx])

                cat_id = row[cat_id_idx] if cat_id_idx != -1 else None
                cat_name = str(row[cat_name_idx]).strip() if cat_name_idx != -1 and row[cat_name_idx] else None
                db_category = self._get_resolved_category(cat_id, cat_name)

                description = row[desc_idx] if desc_idx != -1 else None

                if not name:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Company Expenses", "row": r_num, "error": "Expense Name is required.", "data": str(row)})
                    continue
                if amount is None or amount <= 0:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Company Expenses", "row": r_num, "error": "Amount must be positive.", "data": str(row)})
                    continue
                if not date:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Company Expenses", "row": r_num, "error": "Date is required.", "data": str(row)})
                    continue

                # Check duplicate
                duplicate = CompanyExpense.objects.filter(
                    company_account=db_comp,
                    name=name,
                    amount=amount,
                    date=date,
                    category=db_category
                ).first()

                if duplicate:
                    self.skipped_duplicates += 1
                    if excel_id is not None:
                        self.company_expense_map[excel_id] = duplicate
                    continue

                comp_exp = CompanyExpense.objects.create(
                    company_account=db_comp,
                    name=name,
                    amount=amount,
                    category=db_category,
                    date=date,
                    description=description
                )
                self.imported_counts['company_expenses'] += 1
                if excel_id is not None:
                    self.company_expense_map[excel_id] = comp_exp
            except Exception as e:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Company Expenses", "row": r_num, "error": str(e), "data": str(row)})

    def _import_recurring_transactions(self, wb):
        headers, rows = self._get_sheet_rows(wb, 'recurring_transactions')
        if not headers:
            return

        id_idx = self._find_col_idx(headers, ['id', 'recurring id'])
        name_idx = self._find_col_idx(headers, ['name'])
        type_idx = self._find_col_idx(headers, ['type', 'transaction type'])
        amount_idx = self._find_col_idx(headers, ['amount'])
        cat_id_idx = self._find_col_idx(headers, ['category id'])
        cat_name_idx = self._find_col_idx(headers, ['category name', 'category'])
        acc_id_idx = self._find_col_idx(headers, ['account id'])
        acc_name_idx = self._find_col_idx(headers, ['account name', 'account'])
        comp_id_idx = self._find_col_idx(headers, ['company account id'])
        comp_name_idx = self._find_col_idx(headers, ['company account name', 'company account'])
        freq_idx = self._find_col_idx(headers, ['frequency'])
        start_idx = self._find_col_idx(headers, ['start date'])
        end_idx = self._find_col_idx(headers, ['end date'])
        notes_idx = self._find_col_idx(headers, ['notes', 'note'])
        status_idx = self._find_col_idx(headers, ['status'])

        if name_idx == -1 or type_idx == -1 or amount_idx == -1 or freq_idx == -1 or start_idx == -1:
            return

        for r_num, row in rows:
            excel_id = row[id_idx] if id_idx != -1 else None

            try:
                name = str(row[name_idx]).strip() if row[name_idx] else None
                tx_type = str(row[type_idx]).strip().capitalize() if row[type_idx] else 'Expense'
                if tx_type not in ['Income', 'Expense']:
                    tx_type = 'Expense'

                amount = parse_decimal(row[amount_idx])
                
                cat_id = row[cat_id_idx] if cat_id_idx != -1 else None
                cat_name = str(row[cat_name_idx]).strip() if cat_name_idx != -1 and row[cat_name_idx] else None
                db_category = self._get_resolved_category(cat_id, cat_name)

                if not db_category:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Recurring Transactions", "row": r_num, "error": "Category is required.", "data": str(row)})
                    continue

                acc_id = row[acc_id_idx] if acc_id_idx != -1 else None
                acc_name = str(row[acc_name_idx]).strip() if acc_name_idx != -1 and row[acc_name_idx] else None
                db_account = self._get_resolved_account(acc_id, acc_name) if (acc_id or acc_name) else None

                comp_id = row[comp_id_idx] if comp_id_idx != -1 else None
                comp_name = str(row[comp_name_idx]).strip() if comp_name_idx != -1 and row[comp_name_idx] else None
                db_comp = self._get_resolved_company_account(comp_id, comp_name) if (comp_id or comp_name) else None

                frequency = str(row[freq_idx]).strip().capitalize() if row[freq_idx] else 'Monthly'
                if frequency not in ['Daily', 'Weekly', 'Monthly', 'Quarterly', 'Yearly']:
                    frequency = 'Monthly'

                start_date = parse_date(row[start_idx])
                end_date = parse_date(row[end_idx]) if end_idx != -1 else None
                notes = row[notes_idx] if notes_idx != -1 else None
                status = str(row[status_idx]).strip().capitalize() if status_idx != -1 and row[status_idx] else 'Active'
                if status not in ['Active', 'Inactive']:
                    status = 'Active'

                if not name:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Recurring Transactions", "row": r_num, "error": "Name is required.", "data": str(row)})
                    continue
                if amount is None or amount <= 0:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Recurring Transactions", "row": r_num, "error": "Amount must be positive.", "data": str(row)})
                    continue
                if not start_date:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Recurring Transactions", "row": r_num, "error": "Start date is required.", "data": str(row)})
                    continue

                # Check duplicate
                duplicate = RecurringTransaction.objects.filter(
                    user=self.user,
                    name=name,
                    transaction_type=tx_type,
                    amount=amount,
                    frequency=frequency
                ).first()

                if duplicate:
                    self.skipped_duplicates += 1
                    if excel_id is not None:
                        self.recurring_map[excel_id] = duplicate
                    continue

                recur = RecurringTransaction.objects.create(
                    user=self.user,
                    name=name,
                    transaction_type=tx_type,
                    amount=amount,
                    category=db_category,
                    account=db_account,
                    company_account=db_comp,
                    frequency=frequency,
                    start_date=start_date,
                    end_date=end_date,
                    notes=notes,
                    status=status
                )
                self.imported_counts['recurring_transactions'] += 1
                if excel_id is not None:
                    self.recurring_map[excel_id] = recur
            except Exception as e:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Recurring Transactions", "row": r_num, "error": str(e), "data": str(row)})

    def _import_occurrences(self, wb):
        headers, rows = self._get_sheet_rows(wb, 'generated_occurrences')
        if not headers:
            return

        rec_id_idx = self._find_col_idx(headers, ['recurring transaction id', 'recurring id'])
        rec_name_idx = self._find_col_idx(headers, ['recurring transaction name', 'recurring name', 'recurring'])
        date_idx = self._find_col_idx(headers, ['occurrence date', 'date'])
        inc_id_idx = self._find_col_idx(headers, ['income id'])
        exp_id_idx = self._find_col_idx(headers, ['expense id'])
        cinc_id_idx = self._find_col_idx(headers, ['company income id'])
        cexp_id_idx = self._find_col_idx(headers, ['company expense id'])

        if date_idx == -1:
            return

        for r_num, row in rows:
            try:
                rec_id = row[rec_id_idx] if rec_id_idx != -1 else None
                rec_name = str(row[rec_name_idx]).strip() if rec_name_idx != -1 and row[rec_name_idx] else None
                
                db_recur = None
                if rec_id is not None and rec_id in self.recurring_map:
                    db_recur = self.recurring_map[rec_id]
                elif rec_name:
                    db_recur = RecurringTransaction.objects.filter(user=self.user, name=rec_name).first()

                if not db_recur:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Occurrences", "row": r_num, "error": f"Recurring transaction reference not found (ID: {rec_id}, Name: {rec_name}).", "data": str(row)})
                    continue

                occurrence_date = parse_date(row[date_idx])
                if not occurrence_date:
                    self.skipped_invalid += 1
                    self.errors.append({"sheet": "Occurrences", "row": r_num, "error": "Occurrence date is required.", "data": str(row)})
                    continue

                # Resolve linked transactions if present
                inc_id = row[inc_id_idx] if inc_id_idx != -1 else None
                exp_id = row[exp_id_idx] if exp_id_idx != -1 else None
                cinc_id = row[cinc_id_idx] if cinc_id_idx != -1 else None
                cexp_id = row[cexp_id_idx] if cexp_id_idx != -1 else None

                db_income = self.income_map.get(inc_id) if inc_id is not None else None
                db_expense = self.expense_map.get(exp_id) if exp_id is not None else None
                db_company_income = self.company_income_map.get(cinc_id) if cinc_id is not None else None
                db_company_expense = self.company_expense_map.get(cexp_id) if cexp_id is not None else None

                # Check duplicate
                duplicate = GeneratedOccurrence.objects.filter(
                    recurring_transaction=db_recur,
                    occurrence_date=occurrence_date
                ).exists()

                if duplicate:
                    self.skipped_duplicates += 1
                    continue

                GeneratedOccurrence.objects.create(
                    recurring_transaction=db_recur,
                    occurrence_date=occurrence_date,
                    income=db_income,
                    expense=db_expense,
                    company_income=db_company_income,
                    company_expense=db_company_expense
                )
                self.imported_counts['generated_occurrences'] += 1
            except Exception as e:
                self.skipped_invalid += 1
                self.errors.append({"sheet": "Occurrences", "row": r_num, "error": str(e), "data": str(row)})
