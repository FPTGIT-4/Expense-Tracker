import datetime
from django.views import View
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, Http404
from django.shortcuts import render, redirect
from django.contrib import messages
from django.urls import reverse
from django.utils import timezone

from .exporter import DataExporter
from .importer import DataImporter

class MainIOView(LoginRequiredMixin, TemplateView):
    template_name = 'data_io/main.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Prefill session variables for import report if they exist
        context['import_report'] = self.request.session.pop('import_report', None)
        context['has_log'] = 'import_error_log' in self.request.session

        # Default configurations
        context['modules'] = [
            {'code': 'categories', 'name': 'Categories'},
            {'code': 'labels', 'name': 'Labels'},
            {'code': 'accounts', 'name': 'Accounts'},
            {'code': 'payment_methods', 'name': 'Payment Methods'},
            {'code': 'income', 'name': 'Income'},
            {'code': 'expenses', 'name': 'Expenses'},
            {'code': 'transfers', 'name': 'Transfers'},
            {'code': 'budgets', 'name': 'Budgets'},
            {'code': 'goals', 'name': 'Goals'},
            {'code': 'debts', 'name': 'Debts'},
            {'code': 'repayments', 'name': 'Repayments'},
            {'code': 'company_accounts', 'name': 'Company Accounts'},
            {'code': 'company_income', 'name': 'Company Income'},
            {'code': 'company_expenses', 'name': 'Company Expenses'},
            {'code': 'recurring_transactions', 'name': 'Recurring Transactions'},
            {'code': 'generated_occurrences', 'name': 'Occurrences'},
            {'code': 'transaction_history', 'name': 'Transaction History'},
            {'code': 'user_settings', 'name': 'User Settings'}
        ]
        
        # Get active settings currency
        try:
            settings = self.request.user.settings
            context['currency_symbol'] = settings.currency
        except Exception:
            context['currency_symbol'] = '₹'
            
        return context

class ExportView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        export_format = request.POST.get('format', 'excel')
        date_range = request.POST.get('date_range', 'all')
        custom_start = request.POST.get('start_date')
        custom_end = request.POST.get('end_date')
        selected_modules = request.POST.getlist('modules')

        options = {
            'include_headers': request.POST.get('include_headers') == 'on',
            'include_empty_sheets': request.POST.get('include_empty_sheets') == 'on',
            'compress_zip': request.POST.get('compress_zip') == 'on'
        }

        # If no modules selected, default to all
        if not selected_modules:
            selected_modules = [
                'income', 'expenses', 'categories', 'labels', 'accounts',
                'payment_methods', 'budgets', 'goals', 'debts', 'repayments',
                'company_accounts', 'company_income', 'company_expenses',
                'recurring_transactions', 'generated_occurrences', 'transfers',
                'transaction_history', 'user_settings'
            ]

        exporter = DataExporter(
            user=request.user,
            modules=selected_modules,
            date_range_option=date_range,
            custom_start=custom_start,
            custom_end=custom_end,
            options=options
        )

        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')

        if export_format == 'csv' or options['compress_zip']:
            # CSV ZIP export
            data = exporter.generate_csv_zip()
            response = HttpResponse(data, content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="expense_tracker_export_{timestamp}.zip"'
            return response
        else:
            # Excel export
            data = exporter.generate_excel()
            response = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="expense_tracker_export_{timestamp}.xlsx"'
            return response

class ImportView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        import_file = request.FILES.get('import_file')
        is_restore = request.POST.get('is_restore') == 'on'

        if not import_file:
            messages.error(request, "Please select an Excel (.xlsx) file to upload.")
            return redirect('data_io:main')

        importer = DataImporter(request.user, import_file, is_restore=is_restore)

        # 1. Structural Dry Run Validation
        is_valid, errors, warnings = importer.dry_run_validate()
        if not is_valid:
            # Structurally invalid
            context = {
                'validation_errors': errors,
                'validation_warnings': warnings,
                'modules': [
                    {'code': 'categories', 'name': 'Categories'},
                    {'code': 'labels', 'name': 'Labels'},
                    {'code': 'accounts', 'name': 'Accounts'},
                    {'code': 'payment_methods', 'name': 'Payment Methods'},
                    {'code': 'income', 'name': 'Income'},
                    {'code': 'expenses', 'name': 'Expenses'},
                    {'code': 'transfers', 'name': 'Transfers'},
                    {'code': 'budgets', 'name': 'Budgets'},
                    {'code': 'goals', 'name': 'Goals'},
                    {'code': 'debts', 'name': 'Debts'},
                    {'code': 'repayments', 'name': 'Repayments'},
                    {'code': 'company_accounts', 'name': 'Company Accounts'},
                    {'code': 'company_income', 'name': 'Company Income'},
                    {'code': 'company_expenses', 'name': 'Company Expenses'},
                    {'code': 'recurring_transactions', 'name': 'Recurring Transactions'},
                    {'code': 'generated_occurrences', 'name': 'Occurrences'},
                    {'code': 'transaction_history', 'name': 'Transaction History'},
                    {'code': 'user_settings', 'name': 'User Settings'}
                ]
            }
            try:
                context['currency_symbol'] = request.user.settings.currency
            except Exception:
                context['currency_symbol'] = '₹'
            messages.error(request, "File validation failed. Please check the structural issues listed below.")
            return render(request, 'data_io/main.html', context)

        # 2. Run Import/Restore inside atomic transaction
        import_success = importer.import_data()

        # Compile statistics report
        report = {
            'imported': importer.imported_counts,
            'skipped_duplicates': importer.skipped_duplicates,
            'skipped_invalid': importer.skipped_invalid,
            'success': import_success,
            'is_restore': is_restore
        }

        # If there were row errors, write log to session
        if importer.errors:
            log_lines = []
            log_lines.append("="*60)
            log_lines.append(f"Personal Tracker IMPORT LOG - {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}")
            log_lines.append(f"User: {request.user.username}")
            log_lines.append(f"Mode: {'Full Restore' if is_restore else 'Standard Import'}")
            log_lines.append("="*60)
            log_lines.append(f"Imported items summary: {importer.imported_counts}")
            log_lines.append(f"Skipped duplicates: {importer.skipped_duplicates}")
            log_lines.append(f"Skipped invalid rows: {importer.skipped_invalid}")
            log_lines.append("="*60)
            log_lines.append("\nDetailed errors/warnings:")
            for err in importer.errors:
                log_lines.append(f"[{err['sheet']}][Row {err['row']}] Error: {err['error']}")
                log_lines.append(f"  Data: {err['data']}\n")
            
            request.session['import_error_log'] = '\n'.join(log_lines)
        else:
            # Clear previous log if successful
            request.session.pop('import_error_log', None)

        request.session['import_report'] = report

        if import_success:
            messages.success(request, "Data import completed successfully with no errors!")
        else:
            messages.warning(request, f"Data import completed with some warnings/errors. {importer.skipped_invalid} rows skipped. Download the error report for details.")

        return redirect('data_io:main')

class DownloadImportLogView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        log_content = request.session.get('import_error_log')
        if not log_content:
            messages.info(request, "No import logs available to download.")
            return redirect('data_io:main')

        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        response = HttpResponse(log_content, content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="import_error_log_{timestamp}.txt"'
        return response

class DownloadSampleTemplateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Colors and styles
        banner_font = Font(name='Outfit', size=16, bold=True, color='FFFFFF')
        banner_fill = PatternFill(start_color='1E1B4B', end_color='1E1B4B', fill_type='solid')
        header_font = Font(name='Outfit', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        data_font = Font(name='Inter', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        # Excel sheets specs
        specs = {
            'Categories': {
                'headers': ['Category ID', 'Name', 'Description'],
                'samples': [
                    [1, 'Food', 'All groceries, food delivery, and dining out'],
                    [2, 'Entertainment', 'Movies, music streaming services, gaming, and concerts'],
                    [3, 'Utilities', 'Electricity, water, gas, internet bills, and mobile recharges']
                ]
            },
            'Labels': {
                'headers': ['Label ID', 'Name', 'Color'],
                'samples': [
                    [1, 'Personal', '#10b981'],
                    [2, 'Work', '#ef4444'],
                    [3, 'Urgent', '#f59e0b']
                ]
            },
            'Accounts': {
                'headers': ['Account ID', 'Name', 'Account Type', 'Initial Balance', 'Minimum Balance', 'Status', 'Notes'],
                'samples': [
                    [1, 'Cash Wallet', 'Cash', 500.00, 0.00, 'ACTIVE', 'Physical pocket cash'],
                    [2, 'Chase Savings', 'Bank Account', 15000.00, 1000.00, 'ACTIVE', 'Primary savings account'],
                    [3, 'Amex Card', 'Credit Card', 0.00, 0.00, 'ACTIVE', 'Credit line limit $5000']
                ]
            },
            'Income': {
                'headers': ['Income ID', 'Amount', 'Source', 'Date', 'Account ID', 'Account Name', 'Labels (Comma Separated)', 'Description'],
                'samples': [
                    [1, 5000.00, 'Salary', '2026-06-01', 2, 'Chase Savings', 'Work', 'Monthly corporate pay'],
                    [2, 350.00, 'Freelancing', '2026-06-10', 1, 'Cash Wallet', 'Personal, Urgent', 'Website design gig']
                ]
            },
            'Expenses': {
                'headers': ['Expense ID', 'Name', 'Amount', 'Date', 'Account ID', 'Account Name', 'Category ID', 'Category Name', 'Labels (Comma Separated)', 'Description'],
                'samples': [
                    [1, 'Walmart Groceries', 125.50, '2026-06-02', 2, 'Chase Savings', 1, 'Food', 'Personal', 'Weekly groceries replenishment'],
                    [2, 'Netflix Subscription', 15.99, '2026-06-05', 3, 'Amex Card', 2, 'Entertainment', 'Personal', 'Premium 4K plan subscription']
                ]
            },
            'Transfers': {
                'headers': ['Transfer ID', 'Amount', 'From Account ID', 'From Account Name', 'To Account ID', 'To Account Name', 'Date', 'Note'],
                'samples': [
                    [1, 200.00, 2, 'Chase Savings', 1, 'Cash Wallet', '2026-06-04', 'ATM withdrawal for cash wallet']
                ]
            },
            'Budgets': {
                'headers': ['Budget ID', 'Category ID', 'Category Name', 'Budget Amount', 'Month', 'Year', 'Notes', 'Is Active'],
                'samples': [
                    [1, 1, 'Food', 500.00, 6, 2026, 'Food budget for June', 'True'],
                    [2, 2, 'Entertainment', 100.00, 6, 2026, 'Streaming and hobbies cap', 'True']
                ]
            },
            'Goals': {
                'headers': ['Goal ID', 'Name', 'Target Amount', 'Current Amount', 'Target Date', 'Description'],
                'samples': [
                    [1, 'New iPhone', 1200.00, 300.00, '2026-12-25', 'Upgrade phone by end of year'],
                    [2, 'Emergency Fund', 10000.00, 5000.00, '', '6 months living expenses reserve']
                ]
            },
            'Debts': {
                'headers': ['Debt ID', 'Person Name', 'Debt Type', 'Amount', 'Date', 'Due Date', 'Notes', 'Status'],
                'samples': [
                    [1, 'John Doe', 'Borrowed', 500.00, '2026-05-15', '2026-07-15', 'Rent borrow support', 'Active'],
                    [2, 'Alice Smith', 'Lent', 250.00, '2026-06-01', '2026-06-30', 'Ticket split loan', 'Active']
                ]
            },
            'Repayments': {
                'headers': ['Repayment ID', 'Debt ID', 'Debt Person Name', 'Amount', 'Date', 'Notes'],
                'samples': [
                    [1, 1, 'John Doe', 200.00, '2026-06-01', 'First cash repayment installment']
                ]
            },
            'Company Accounts': {
                'headers': ['Company Account ID', 'Name', 'Description', 'Opening Balance', 'Created Date', 'Status'],
                'samples': [
                    [1, 'Acme Corp Fin', 'Acme consulting operations account', 5000.00, '2026-01-01', 'ACTIVE']
                ]
            },
            'Company Income': {
                'headers': ['Company Income ID', 'Company Account ID', 'Company Account Name', 'Amount', 'Source', 'Date', 'Description'],
                'samples': [
                    [1, 1, 'Acme Corp Fin', 3500.00, 'Client Retainer', '2026-06-01', 'Monthly client retainer payment']
                ]
            },
            'Company Expenses': {
                'headers': ['Company Expense ID', 'Company Account ID', 'Company Account Name', 'Name', 'Amount', 'Category ID', 'Category Name', 'Date', 'Description'],
                'samples': [
                    [1, 1, 'Acme Corp Fin', 'AWS Web Hosting', 85.00, 3, 'Utilities', '2026-06-02', 'Cloud infrastructure billing']
                ]
            },
            'Recurring Transactions': {
                'headers': ['Recurring ID', 'Name', 'Type', 'Amount', 'Category ID', 'Category Name', 'Account ID', 'Account Name', 'Company Account ID', 'Company Account Name', 'Frequency', 'Start Date', 'End Date', 'Notes', 'Status'],
                'samples': [
                    [1, 'Gym Membership', 'Expense', 50.00, 2, 'Entertainment', 3, 'Amex Card', '', '', 'Monthly', '2026-01-01', '', 'Monthly fitness sub', 'Active']
                ]
            },
            'Occurrences': {
                'headers': ['Occurrence ID', 'Recurring ID', 'Recurring Name', 'Occurrence Date', 'Income ID', 'Expense ID', 'Company Income ID', 'Company Expense ID'],
                'samples': [
                    [1, 1, 'Gym Membership', '2026-06-01', '', 2, '', '']
                ]
            },
            'Transaction History': {
                'headers': ['History ID', 'Activity Type', 'Account ID', 'Account Name', 'To Account ID', 'To Account Name', 'Amount', 'Balance Before', 'Balance After', 'Category Name', 'Description', 'Date', 'Timestamp'],
                'samples': [
                    [1, 'INCOME', 1, 'Checking Account', '', '', 3500.00, 0.00, 3500.00, 'Salary', 'Monthly software engineer salary', '2026-04-01', '2026-04-01 09:00:00']
                ]
            },
            'User Settings': {
                'headers': ['Setting ID', 'Currency', 'Budget Threshold', 'Enable Budget Alerts', 'Low Balance Alerts', 'Show Navbar Badge', 'Show Dashboard Banner', 'Show Dashboard Panel', 'Alert Scope', 'Default Minimum Balance', 'Dark Mode'],
                'samples': [
                    [1, '₹', 80, 'True', 'True', 'True', 'True', 'True', 'active', 0.00, 'True']
                ]
            }
        }

        for sheet_name, spec in specs.items():
            ws = wb.create_sheet(title=sheet_name)
            ws.views.sheetView[0].showGridLines = True

            # Banner row 1
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(spec['headers']))
            banner = ws.cell(row=1, column=1)
            banner.value = f"  {sheet_name} Sample Template (Row 4 is example data. Delete before uploading!)"
            banner.font = banner_font
            banner.fill = banner_fill
            banner.alignment = Alignment(vertical='center')
            ws.row_dimensions[1].height = 40

            # Header row 3
            for col_idx, header in enumerate(spec['headers'], 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            ws.row_dimensions[3].height = 25

            # Sample row 4
            for s_row_idx, s_row in enumerate(spec['samples'], 4):
                for col_idx, val in enumerate(s_row, 1):
                    cell = ws.cell(row=s_row_idx, column=col_idx)
                    cell.value = val
                    cell.font = data_font
                    cell.alignment = left_align
                    if isinstance(val, (int, float)):
                        cell.number_format = '0.00'
                ws.row_dimensions[s_row_idx].height = 20

            # Column widths
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    if cell.row == 1:
                        continue
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="expense_tracker_import_template.xlsx"'
        wb.save(response)
        return response
