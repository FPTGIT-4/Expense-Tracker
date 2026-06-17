import io
import csv
import zipfile
import datetime
from decimal import Decimal
from django.utils import timezone
from django.db.models import Q

# Import all models
from income.models import Income
from expenses.models import Expense
from categories.models import Category, Label
from accounts.models import Account, AccountTransfer, UserSettings
from budgets.models import Budget
from goals.models import Goal
from debts.models import Debt, Repayment
from companies.models import CompanyAccount, CompanyIncome, CompanyExpense
from recurrences.models import RecurringTransaction, GeneratedOccurrence
from transactions.models import TransactionHistory

def get_date_range(filter_option, custom_start=None, custom_end=None):
    today = timezone.localdate()
    start_date = None
    end_date = None

    if filter_option == 'today':
        start_date = today
        end_date = today
    elif filter_option == 'yesterday':
        start_date = today - datetime.timedelta(days=1)
        end_date = today - datetime.timedelta(days=1)
    elif filter_option == 'this_week':
        start_date = today - datetime.timedelta(days=today.weekday())
        end_date = today
    elif filter_option == 'last_week':
        start_week = today - datetime.timedelta(days=today.weekday() + 7)
        end_date = start_week + datetime.timedelta(days=6)
        start_date = start_week
    elif filter_option == 'this_month':
        start_date = today.replace(day=1)
        end_date = today
    elif filter_option == 'last_month':
        first_day_this_month = today.replace(day=1)
        last_day_last_month = first_day_this_month - datetime.timedelta(days=1)
        start_date = last_day_last_month.replace(day=1)
        end_date = last_day_last_month
    elif filter_option == 'this_year':
        start_date = today.replace(month=1, day=1)
        end_date = today
    elif filter_option == 'last_year':
        start_date = datetime.date(today.year - 1, 1, 1)
        end_date = datetime.date(today.year - 1, 12, 31)
    elif filter_option == 'custom' and custom_start and custom_end:
        if isinstance(custom_start, str):
            try:
                start_date = datetime.datetime.strptime(custom_start, '%Y-%m-%d').date()
            except ValueError:
                pass
        else:
            start_date = custom_start
        if isinstance(custom_end, str):
            try:
                end_date = datetime.datetime.strptime(custom_end, '%Y-%m-%d').date()
            except ValueError:
                pass
        else:
            end_date = custom_end

    return start_date, end_date

class DataExporter:
    def __init__(self, user, modules=None, date_range_option='all', custom_start=None, custom_end=None, options=None):
        self.user = user
        self.modules = modules or [
            'income', 'expenses', 'categories', 'labels', 'accounts',
            'payment_methods', 'budgets', 'goals', 'debts', 'repayments',
            'company_accounts', 'company_income', 'company_expenses',
            'recurring_transactions', 'generated_occurrences', 'transfers',
            'transaction_history', 'user_settings'
        ]
        self.date_range_option = date_range_option
        self.start_date, self.end_date = get_date_range(date_range_option, custom_start, custom_end)
        self.options = options or {
            'include_headers': True,
            'include_empty_sheets': True,
            'compress_zip': False
        }

    def _get_queryset(self, model_name):
        """Helper to get user-specific filtered querysets."""
        if model_name == 'categories':
            return Category.objects.filter(user=self.user)
        elif model_name == 'labels':
            return Label.objects.filter(user=self.user)
        elif model_name == 'accounts' or model_name == 'payment_methods':
            return Account.objects.filter(user=self.user)
        elif model_name == 'income':
            qs = Income.objects.filter(user=self.user)
            if self.start_date and self.end_date:
                qs = qs.filter(date__range=(self.start_date, self.end_date))
            return qs
        elif model_name == 'expenses':
            qs = Expense.objects.filter(user=self.user)
            if self.start_date and self.end_date:
                qs = qs.filter(date__range=(self.start_date, self.end_date))
            return qs
        elif model_name == 'transfers':
            qs = AccountTransfer.objects.filter(user=self.user)
            if self.start_date and self.end_date:
                qs = qs.filter(transfer_date__range=(self.start_date, self.end_date))
            return qs
        elif model_name == 'budgets':
            qs = Budget.objects.filter(user=self.user)
            if self.start_date and self.end_date:
                # Filter budgets active or defined within year/month of range
                qs = qs.filter(
                    Q(year__gt=self.start_date.year) | Q(year=self.start_date.year, month__gte=self.start_date.month)
                ).filter(
                    Q(year__lt=self.end_date.year) | Q(year=self.end_date.year, month__lte=self.end_date.month)
                )
            return qs
        elif model_name == 'goals':
            qs = Goal.objects.filter(user=self.user)
            if self.start_date and self.end_date:
                qs = qs.filter(Q(target_date__range=(self.start_date, self.end_date)) | Q(target_date__isnull=True))
            return qs
        elif model_name == 'debts':
            qs = Debt.objects.filter(user=self.user)
            if self.start_date and self.end_date:
                qs = qs.filter(date__range=(self.start_date, self.end_date))
            return qs
        elif model_name == 'repayments':
            qs = Repayment.objects.filter(debt__user=self.user)
            if self.start_date and self.end_date:
                qs = qs.filter(date__range=(self.start_date, self.end_date))
            return qs
        elif model_name == 'company_accounts':
            return CompanyAccount.objects.filter(user=self.user)
        elif model_name == 'company_income':
            qs = CompanyIncome.objects.filter(company_account__user=self.user)
            if self.start_date and self.end_date:
                qs = qs.filter(date__range=(self.start_date, self.end_date))
            return qs
        elif model_name == 'company_expenses':
            qs = CompanyExpense.objects.filter(company_account__user=self.user)
            if self.start_date and self.end_date:
                qs = qs.filter(date__range=(self.start_date, self.end_date))
            return qs
        elif model_name == 'recurring_transactions':
            return RecurringTransaction.objects.filter(user=self.user)
        elif model_name == 'generated_occurrences':
            qs = GeneratedOccurrence.objects.filter(recurring_transaction__user=self.user)
            if self.start_date and self.end_date:
                qs = qs.filter(occurrence_date__range=(self.start_date, self.end_date))
            return qs
        elif model_name == 'transaction_history':
            qs = TransactionHistory.objects.filter(user=self.user)
            if self.start_date and self.end_date:
                qs = qs.filter(date__range=(self.start_date, self.end_date))
            return qs
        elif model_name == 'user_settings':
            return UserSettings.objects.filter(user=self.user)
        return None

    def _get_sheet_specs(self):
        """Returns the sheet definition, headers, and row serialization functions."""
        return {
            'categories': {
                'title': 'Categories',
                'headers': ['Category ID', 'Name', 'Description', 'Created At', 'Updated At'],
                'data_func': lambda obj: [
                    obj.id, obj.name, obj.description,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None,
                    obj.updated_at.replace(tzinfo=None) if obj.updated_at else None
                ]
            },
            'labels': {
                'title': 'Labels',
                'headers': ['Label ID', 'Name', 'Color', 'Created At', 'Updated At'],
                'data_func': lambda obj: [
                    obj.id, obj.name, obj.color,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None,
                    obj.updated_at.replace(tzinfo=None) if obj.updated_at else None
                ]
            },
            'accounts': {
                'title': 'Accounts',
                'headers': ['Account ID', 'Name', 'Account Type', 'Initial Balance', 'Minimum Balance', 'Status', 'Notes', 'Created At', 'Updated At'],
                'data_func': lambda obj: [
                    obj.id, obj.name, obj.account_type, obj.initial_balance, obj.minimum_balance, obj.status, obj.notes,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None,
                    obj.updated_at.replace(tzinfo=None) if obj.updated_at else None
                ]
            },
            'payment_methods': {
                'title': 'Payment Methods',
                'headers': ['Payment Method ID', 'Name', 'Account Type', 'Initial Balance', 'Minimum Balance', 'Status', 'Notes', 'Created At', 'Updated At'],
                'data_func': lambda obj: [
                    obj.id, obj.name, obj.account_type, obj.initial_balance, obj.minimum_balance, obj.status, obj.notes,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None,
                    obj.updated_at.replace(tzinfo=None) if obj.updated_at else None
                ]
            },
            'income': {
                'title': 'Income',
                'headers': ['Income ID', 'Amount', 'Source', 'Date', 'Account ID', 'Account Name', 'Labels (Comma Separated)', 'Description', 'Created At', 'Updated At'],
                'data_func': lambda obj: [
                    obj.id, obj.amount, obj.source, obj.date,
                    obj.account.id if obj.account else None,
                    obj.account.name if obj.account else None,
                    ', '.join([l.name for l in obj.labels.all()]),
                    obj.description,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None,
                    obj.updated_at.replace(tzinfo=None) if obj.updated_at else None
                ]
            },
            'expenses': {
                'title': 'Expenses',
                'headers': ['Expense ID', 'Name', 'Amount', 'Date', 'Account ID', 'Account Name', 'Category ID', 'Category Name', 'Labels (Comma Separated)', 'Description', 'Created At', 'Updated At'],
                'data_func': lambda obj: [
                    obj.id, obj.name, obj.amount, obj.date,
                    obj.account.id if obj.account else None,
                    obj.account.name if obj.account else None,
                    obj.category.id if obj.category else None,
                    obj.category.name if obj.category else None,
                    ', '.join([l.name for l in obj.labels.all()]),
                    obj.description,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None,
                    obj.updated_at.replace(tzinfo=None) if obj.updated_at else None
                ]
            },
            'transfers': {
                'title': 'Transfers',
                'headers': ['Transfer ID', 'Amount', 'From Account ID', 'From Account Name', 'To Account ID', 'To Account Name', 'Date', 'Note', 'Created At'],
                'data_func': lambda obj: [
                    obj.id, obj.amount,
                    obj.from_account.id if obj.from_account else None,
                    obj.from_account.name if obj.from_account else None,
                    obj.to_account.id if obj.to_account else None,
                    obj.to_account.name if obj.to_account else None,
                    obj.transfer_date, obj.note,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None
                ]
            },
            'budgets': {
                'title': 'Budgets',
                'headers': ['Budget ID', 'Category ID', 'Category Name', 'Budget Amount', 'Month', 'Year', 'Notes', 'Is Active', 'Created At', 'Updated At'],
                'data_func': lambda obj: [
                    obj.id, obj.category.id, obj.category.name,
                    obj.budget_amount, obj.month, obj.year, obj.notes, obj.is_active,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None,
                    obj.updated_at.replace(tzinfo=None) if obj.updated_at else None
                ]
            },
            'goals': {
                'title': 'Goals',
                'headers': ['Goal ID', 'Name', 'Target Amount', 'Current Amount', 'Target Date', 'Description', 'Created At', 'Updated At'],
                'data_func': lambda obj: [
                    obj.id, obj.name, obj.target_amount, obj.current_amount, obj.target_date, obj.description,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None,
                    obj.updated_at.replace(tzinfo=None) if obj.updated_at else None
                ]
            },
            'debts': {
                'title': 'Debts',
                'headers': ['Debt ID', 'Person Name', 'Debt Type', 'Amount', 'Date', 'Due Date', 'Notes', 'Status', 'Created At', 'Updated At'],
                'data_func': lambda obj: [
                    obj.id, obj.person_name, obj.debt_type, obj.amount, obj.date, obj.due_date, obj.notes, obj.status,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None,
                    obj.updated_at.replace(tzinfo=None) if obj.updated_at else None
                ]
            },
            'repayments': {
                'title': 'Repayments',
                'headers': ['Repayment ID', 'Debt ID', 'Debt Person Name', 'Amount', 'Date', 'Notes', 'Created At', 'Updated At'],
                'data_func': lambda obj: [
                    obj.id, obj.debt.id, obj.debt.person_name,
                    obj.amount, obj.date, obj.notes,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None,
                    obj.updated_at.replace(tzinfo=None) if obj.updated_at else None
                ]
            },
            'company_accounts': {
                'title': 'Company Accounts',
                'headers': ['Company Account ID', 'Name', 'Description', 'Opening Balance', 'Created Date', 'Status', 'Created At', 'Updated At'],
                'data_func': lambda obj: [
                    obj.id, obj.name, obj.description, obj.opening_balance, obj.created_date, obj.status,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None,
                    obj.updated_at.replace(tzinfo=None) if obj.updated_at else None
                ]
            },
            'company_income': {
                'title': 'Company Income',
                'headers': ['Company Income ID', 'Company Account ID', 'Company Account Name', 'Amount', 'Source', 'Date', 'Description', 'Created At', 'Updated At'],
                'data_func': lambda obj: [
                    obj.id, obj.company_account.id, obj.company_account.name,
                    obj.amount, obj.source, obj.date, obj.description,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None,
                    obj.updated_at.replace(tzinfo=None) if obj.updated_at else None
                ]
            },
            'company_expenses': {
                'title': 'Company Expenses',
                'headers': ['Company Expense ID', 'Company Account ID', 'Company Account Name', 'Name', 'Amount', 'Category ID', 'Category Name', 'Date', 'Description', 'Created At', 'Updated At'],
                'data_func': lambda obj: [
                    obj.id, obj.company_account.id, obj.company_account.name,
                    obj.name, obj.amount,
                    obj.category.id if obj.category else None,
                    obj.category.name if obj.category else None,
                    obj.date, obj.description,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None,
                    obj.updated_at.replace(tzinfo=None) if obj.updated_at else None
                ]
            },
            'recurring_transactions': {
                'title': 'Recurring Transactions',
                'headers': ['Recurring ID', 'Name', 'Type', 'Amount', 'Category ID', 'Category Name', 'Account ID', 'Account Name', 'Company Account ID', 'Company Account Name', 'Frequency', 'Start Date', 'End Date', 'Notes', 'Status', 'Created At', 'Updated At'],
                'data_func': lambda obj: [
                    obj.id, obj.name, obj.transaction_type, obj.amount,
                    obj.category.id, obj.category.name,
                    obj.account.id if obj.account else None,
                    obj.account.name if obj.account else None,
                    obj.company_account.id if obj.company_account else None,
                    obj.company_account.name if obj.company_account else None,
                    obj.frequency, obj.start_date, obj.end_date, obj.notes, obj.status,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None,
                    obj.updated_at.replace(tzinfo=None) if obj.updated_at else None
                ]
            },
            'generated_occurrences': {
                'title': 'Occurrences',
                'headers': ['Occurrence ID', 'Recurring ID', 'Recurring Name', 'Occurrence Date', 'Income ID', 'Expense ID', 'Company Income ID', 'Company Expense ID', 'Created At'],
                'data_func': lambda obj: [
                    obj.id, obj.recurring_transaction.id, obj.recurring_transaction.name,
                    obj.occurrence_date,
                    obj.income.id if obj.income else None,
                    obj.expense.id if obj.expense else None,
                    obj.company_income.id if obj.company_income else None,
                    obj.company_expense.id if obj.company_expense else None,
                    obj.created_at.replace(tzinfo=None) if obj.created_at else None
                ]
            },
            'transaction_history': {
                'title': 'Transaction History',
                'headers': ['History ID', 'Activity Type', 'Account ID', 'Account Name', 'To Account ID', 'To Account Name', 'Amount', 'Balance Before', 'Balance After', 'Category Name', 'Description', 'Date', 'Timestamp'],
                'data_func': lambda obj: [
                    obj.id, obj.activity_type,
                    obj.account.id if obj.account else None,
                    obj.account.name if obj.account else None,
                    obj.to_account.id if obj.to_account else None,
                    obj.to_account.name if obj.to_account else None,
                    obj.amount, obj.balance_before, obj.balance_after,
                    obj.category_name, obj.description, obj.date,
                    obj.timestamp.replace(tzinfo=None) if obj.timestamp else None
                ]
            },
            'user_settings': {
                'title': 'User Settings',
                'headers': ['Setting ID', 'Currency', 'Budget Threshold', 'Enable Budget Alerts', 'Low Balance Alerts', 'Show Navbar Badge', 'Show Dashboard Banner', 'Show Dashboard Panel', 'Alert Scope', 'Default Minimum Balance', 'Dark Mode'],
                'data_func': lambda obj: [
                    obj.id, obj.currency, obj.budget_threshold,
                    obj.enable_budget_alerts, obj.low_balance_alerts,
                    obj.low_balance_show_navbar_badge, obj.low_balance_show_dashboard_banner,
                    obj.low_balance_show_dashboard_panel, obj.low_balance_alert_scope,
                    obj.low_balance_default_minimum, obj.dark_mode
                ]
            }
        }

    def generate_excel(self):
        """Generates an in-memory openpyxl Excel file based on options."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        sheet_specs = self._get_sheet_specs()

        # Excel styles
        title_font = Font(name='Outfit', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='1E1B4B', end_color='1E1B4B', fill_type='solid') # Slate indigo
        header_font = Font(name='Outfit', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid') # Indigo
        data_font = Font(name='Inter', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        for mod in self.modules:
            spec = sheet_specs.get(mod)
            if not spec:
                continue

            qs = self._get_queryset(mod)
            if qs is None:
                continue

            has_data = qs.exists()
            if not has_data and not self.options.get('include_empty_sheets', True):
                continue

            ws = wb.create_sheet(title=spec['title'])
            ws.views.sheetView[0].showGridLines = True

            row_idx = 1
            # Add Sheet Title Banner
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(spec['headers']))
            title_cell = ws.cell(row=1, column=1)
            title_cell.value = f"  {spec['title']} Report"
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = Alignment(vertical='center')
            ws.row_dimensions[1].height = 40
            row_idx += 1 # Empty row
            row_idx += 1 # Header row

            # Write Headers
            if self.options.get('include_headers', True):
                for col_idx, header in enumerate(spec['headers'], 1):
                    cell = ws.cell(row=3, column=col_idx)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                ws.row_dimensions[3].height = 25
                row_idx = 4

            # Write Data rows
            for obj in qs:
                row_data = spec['data_func'](obj)
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = val
                    cell.font = data_font
                    cell.alignment = left_align
                    
                    # Number format conversions
                    if isinstance(val, (Decimal, float)):
                        cell.number_format = '#,##0.00'
                    elif isinstance(val, (datetime.date, datetime.datetime)):
                        cell.number_format = 'yyyy-mm-dd'
                ws.row_dimensions[row_idx].height = 20
                row_idx += 1

            # Auto-fit columns
            for col in ws.columns:
                max_len = 0
                for cell in col:
                    if cell.row == 1:
                        continue # Skip merged title row when calculating width
                    if cell.value:
                        if isinstance(cell.value, datetime.date):
                            max_len = max(max_len, 10)
                        elif isinstance(cell.value, (Decimal, float)):
                            max_len = max(max_len, len(f"{cell.value:.2f}"))
                        else:
                            max_len = max(max_len, len(str(cell.value)))
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_csv_zip(self):
        """Generates an in-memory ZIP containing separate CSV files for each module."""
        sheet_specs = self._get_sheet_specs()
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for mod in self.modules:
                spec = sheet_specs.get(mod)
                if not spec:
                    continue

                qs = self._get_queryset(mod)
                if qs is None:
                    continue

                has_data = qs.exists()
                if not has_data and not self.options.get('include_empty_sheets', True):
                    continue

                csv_buffer = io.StringIO()
                writer = csv.writer(csv_buffer)

                # Headers
                if self.options.get('include_headers', True):
                    writer.writerow(spec['headers'])

                # Data rows
                for obj in qs:
                    row_data = spec['data_func'](obj)
                    # format values as strings for CSV compatibility
                    formatted_row = []
                    for val in row_data:
                        if isinstance(val, (datetime.date, datetime.datetime)):
                            formatted_row.append(val.strftime('%Y-%m-%d'))
                        elif isinstance(val, Decimal):
                            formatted_row.append(f"{val:.2f}")
                        else:
                            formatted_row.append(str(val) if val is not None else '')
                    writer.writerow(formatted_row)

                file_name = f"{spec['title'].lower().replace(' ', '_')}.csv"
                zip_file.writestr(file_name, csv_buffer.getvalue())

        zip_buffer.seek(0)
        return zip_buffer.getvalue()
