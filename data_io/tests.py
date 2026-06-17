import io
import datetime
from decimal import Decimal
import openpyxl
from django.test import TestCase, Client
from django.contrib.auth.models import User
from django.urls import reverse
from django.utils import timezone

from categories.models import Category
from accounts.models import Account
from income.models import Income
from expenses.models import Expense
from data_io.exporter import DataExporter
from data_io.importer import DataImporter

class DataIOTestCase(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='testuser', password='password123')
        self.client = Client()
        self.client.login(username='testuser', password='password123')

        # Create some master data
        self.cat_food = Category.objects.create(user=self.user, name='Food', description='Eating expenses')
        self.acc_cash = Account.objects.create(user=self.user, name='Pocket Cash', account_type='Cash', initial_balance=Decimal('100.00'))

        # Create some transactions
        self.inc = Income.objects.create(user=self.user, account=self.acc_cash, amount=Decimal('50.00'), source='Freelancing', date=timezone.localdate())
        self.exp = Expense.objects.create(user=self.user, account=self.acc_cash, amount=Decimal('20.00'), name='Lunch', date=timezone.localdate(), category=self.cat_food)

    def test_main_page_renders(self):
        response = self.client.get(reverse('data_io:main'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Manage your financial data portability")

    def test_export_excel_generation(self):
        exporter = DataExporter(user=self.user)
        excel_data = exporter.generate_excel()
        self.assertIsInstance(excel_data, bytes)
        
        # Load exported file to verify sheets and rows
        wb = openpyxl.load_workbook(io.BytesIO(excel_data))
        self.assertIn('Categories', wb.sheetnames)
        self.assertIn('Accounts', wb.sheetnames)
        self.assertIn('Income', wb.sheetnames)
        self.assertIn('Expenses', wb.sheetnames)

        # Check Category sheet values
        cat_ws = wb['Categories']
        # Banner row 1, empty row 2, headers row 3, data row 4
        self.assertEqual(cat_ws.cell(row=3, column=2).value, 'Name')
        self.assertEqual(cat_ws.cell(row=4, column=2).value, 'Food')

    def test_export_view_returns_file(self):
        response = self.client.post(reverse('data_io:export'), {
            'format': 'excel',
            'date_range': 'all',
            'modules': ['categories', 'accounts']
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_sample_template_generation(self):
        response = self.client.get(reverse('data_io:sample_template'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertIn('Categories', wb.sheetnames)
        self.assertIn('Income', wb.sheetnames)

    def test_importer_dry_run_validation(self):
        # Create a valid in-memory workbook
        wb = openpyxl.Workbook()
        ws_cat = wb.active
        ws_cat.title = "Categories"
        ws_cat.cell(row=1, column=1, value="Categories Report") # Banner
        ws_cat.cell(row=3, column=1, value="Category ID")
        ws_cat.cell(row=3, column=2, value="Name")
        ws_cat.cell(row=3, column=3, value="Description")
        ws_cat.cell(row=4, column=1, value=10)
        ws_cat.cell(row=4, column=2, value="Transport")
        ws_cat.cell(row=4, column=3, value="Bus fares")

        ws_acc = wb.create_sheet(title="Accounts")
        ws_acc.cell(row=1, column=1, value="Accounts Report")
        ws_acc.cell(row=3, column=1, value="Account ID")
        ws_acc.cell(row=3, column=2, value="Name")
        ws_acc.cell(row=3, column=3, value="Account Type")
        ws_acc.cell(row=4, column=1, value=20)
        ws_acc.cell(row=4, column=2, value="Credit Card")
        ws_acc.cell(row=4, column=3, value="Credit Card")

        ws_inc = wb.create_sheet(title="Income")
        ws_inc.cell(row=3, column=1, value="Income ID")
        ws_inc.cell(row=3, column=2, value="Amount")
        ws_inc.cell(row=3, column=3, value="Source")
        ws_inc.cell(row=3, column=4, value="Date")

        ws_exp = wb.create_sheet(title="Expenses")
        ws_exp.cell(row=3, column=1, value="Expense ID")
        ws_exp.cell(row=3, column=2, value="Name")
        ws_exp.cell(row=3, column=3, value="Amount")
        ws_exp.cell(row=3, column=4, value="Date")

        # Save workbook to buffer
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        importer = DataImporter(self.user, buf)
        is_valid, errors, warnings = importer.dry_run_validate()
        self.assertTrue(is_valid)
        self.assertEqual(len(errors), 0)

        # Execute import
        buf.seek(0)
        success = importer.import_data()
        self.assertTrue(success)

        # Verify that Category 'Transport' was created
        self.assertTrue(Category.objects.filter(user=self.user, name='Transport').exists())
        # Verify that Account 'Credit Card' was created
        self.assertTrue(Account.objects.filter(user=self.user, name='Credit Card').exists())
